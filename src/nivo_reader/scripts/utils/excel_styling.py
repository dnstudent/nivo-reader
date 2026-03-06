#!/usr/bin/env python3
"""
nivo-reader: a tool to automatically read meteorological data
Copyright (C) 2026  Davide Nicoli

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, override

import polars as pl
from openpyxl.styles import Font, PatternFill


class ExcelStyler(ABC):
    """Interface for Excel cell stylers."""

    @abstractmethod
    def apply(
        self,
        ws: Any,
        aggregated_df: pl.DataFrame,
        table_region: tuple[int, int, int, int],
    ) -> None:
        """Apply styles to the worksheet.

        Args:
            ws (Any): The worksheet object.
            aggregated_df (pl.DataFrame): The aggregated data (long format).
            table_region (tuple[int, int, int, int]): The region where data was written (start_row, start_col, end_row, end_col), all 1-indexed.
        """
        pass


class NivoDefaultStyler(ExcelStyler):
    """Default styler for NIVO tables. Highlights cells with low confidence."""

    threshold: float
    low_conf_fill: PatternFill
    low_conf_font: Font
    error_fill: PatternFill
    error_font: Font

    def __init__(
        self,
        threshold: float,
    ):
        """
        Args:
            threshold (float): Confidence threshold below which cells are highlighted.
        """
        self.threshold = threshold
        self.low_conf_fill = PatternFill(
            bgColor="FFFFEB9C",
            fgColor="FFFFEB9C",
            fill_type="solid",
        )  # Light yellow fill (Excel's default for "Warning")
        self.low_conf_font = Font(color="FF9C6500")
        self.error_fill = PatternFill(
            bgColor="FFFFC7CE", fgColor="FFFFC7CE", fill_type="solid"
        )  # Heavy red fill for empty cells
        self.error_font = Font(color="FF9C0006")

    @override
    def apply(
        self,
        ws: Any,
        aggregated_df: pl.DataFrame,
        table_region: tuple[int, int, int, int],
    ) -> None:
        """Style the table based on confidence values."""
        start_row, start_col, _, _ = table_region

        data_lookup = {
            (r, c): {"confidence": conf, "content": content}
            for r, c, conf, content in aggregated_df.select(
                ["row", "column", "confidence", "content"]
            ).iter_rows()
        }

        for (data_row, data_col), ocr_result in data_lookup.items():
            if (
                ocr_result["confidence"] is None
                or ocr_result["content"] == ""
                or ocr_result["content"] is None
            ):
                ws_row = start_row + data_row
                ws_col = start_col + data_col
                ws.cell(row=ws_row, column=ws_col).fill = self.error_fill
                ws.cell(row=ws_row, column=ws_col).font = self.error_font

            elif ocr_result["confidence"] < self.threshold:
                # ws is 1-indexed. data_row/data_col are 0-indexed relative to the table start.
                ws_row = start_row + data_row
                ws_col = start_col + data_col
                ws.cell(row=ws_row, column=ws_col).fill = self.low_conf_fill
                ws.cell(row=ws_row, column=ws_col).font = self.low_conf_font


def apply_styler(
    wb_path: str | Path,
    styler: ExcelStyler,
    aggregated_df: pl.DataFrame,
    table_region: tuple[int, int, int, int],
    sheet_name: str = "Sheet1",
):
    """Apply a styler to a workbook and save it.

    Args:
        wb_path (str | Path): Path to the workbook file.
        styler (ExcelStyler): The styler to apply.
        aggregated_df (pl.DataFrame): The aggregated data.
        table_region (tuple[int, int, int, int]): The region where data was written.
        sheet_name (str): The name of the sheet to style.
    """
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    styler.apply(ws, aggregated_df, table_region)

    wb.save(wb_path)
    wb.close()
