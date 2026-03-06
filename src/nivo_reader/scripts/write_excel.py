#!/usr/bin/env python3
"""
nivo-reader: a tool to digitize snowfall data tables from the Italian Hydrological Service
Copyright (C) 2026  Davide Nicoli

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import argparse
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from nivo_reader.modules.results_aggregation import (
    AggregatorPipeline,
    HighestScoring,
    MostOccurringValues,
)
from nivo_reader.modules.results_aggregation.base import ResultsAggregator

from .utils.excel_styling import ExcelStyler, NivoDefaultStyler
from .utils.paths import discover_files, reroute_file


def create_parser():
    parser = argparse.ArgumentParser(
        prog="nivo-writer",
        description="Batch conversion of NIVO table digitizations to Excel",
    )
    _ = parser.add_argument(
        "input_dir",
        help="Directory containing the NIVO table digitization files",
        type=str,
    )
    _ = parser.add_argument(
        "--output-dir",
        help="Directory where the Excel files will be saved",
        type=str,
        default=None,
    )
    _ = parser.add_argument(
        "--header-template",
        help="Path to a template Excel file to use as a header",
        type=str,
        default=None,
    )
    _ = parser.add_argument(
        "--confidence-threshold",
        help="Confidence threshold for highlighting cells",
        type=float,
        default=None,
    )
    return parser


def write_excel_workbook(
    digitization_df: pl.DataFrame,
    aggregator: ResultsAggregator,
    styler: ExcelStyler,
    header_template: Path | None = None,
) -> Workbook:
    """Create a styled Excel Workbook from a single digitization dataframe.

    Args:
        digitization_df: The input digitization dataframe (raw or cleaned)
        aggregator: The ResultsAggregator pipeline to apply
        styler: The ExcelStyler to apply to the workbook
        header_template: Optional path to an Excel template file to use as base

    Returns:
        The generated and styled openpyxl Workbook
    """
    row = 0
    col = 0
    if header_template and header_template.exists():
        wb = load_workbook(header_template)
        # Type assertion for ws since active can be None or Worksheet, memory-only
        ws = wb.active
        assert isinstance(ws, Worksheet)
        if ws:
            row = ws.max_row
            col = ws.min_column - 1
            ws.title = "Sheet1"
        else:
            ws = wb.create_sheet("Sheet1")
    else:
        wb = Workbook()
        ws = wb.active
        assert isinstance(ws, Worksheet)
        ws.title = "Sheet1"

    aggregated_df = aggregator(digitization_df)

    # We use pandas to write to the worksheet (via openpyxl utilities)
    # openpyxl utils 'dataframe_to_rows' requires pandas or openpyxl native iteration.
    # It's easier to use openpyxl's `utils.dataframe.dataframe_to_rows`
    # However we already used `pd.ExcelWriter` with `mode='a'`.
    # Let's write the dataframe directly using iter_rows or pandas' internal wrapper
    pd_df = (
        aggregated_df.pivot(
            on="column", index="row", values="content", sort_columns=True
        )
        .sort(by="row")
        .drop("row")
        .to_pandas()
    )

    from openpyxl.utils.dataframe import dataframe_to_rows

    # startrow=row, startcol=col in 0-indexed pandas mapping
    for r_idx, pd_row in enumerate(dataframe_to_rows(pd_df, index=False, header=False)):
        for c_idx, value in enumerate(pd_row):
            _ = ws.cell(row=row + r_idx + 1, column=col + c_idx + 1, value=value)

    # Calculate table region (1-indexed for openpyxl)
    num_rows = aggregated_df.get_column("row").n_unique()
    num_cols = aggregated_df.get_column("column").n_unique()
    table_region = (row + 1, col + 1, row + num_rows, col + num_cols)

    # Apply styling
    styler.apply(ws, aggregated_df, table_region)

    return wb


def write_excels_batch(
    input_dir: Path,
    output_dir: Path,
    aggregator: ResultsAggregator,
    styler: ExcelStyler,
    header_template: Path | None = None,
) -> None:
    """Process multiple digitizations and save styled Excel templates.

    Args:
        input_dir: Directory containing digitization files
        output_dir: Directory where Excel files will be saved
        aggregator: Aggregator logic to condense the OCR readings
        styler: Styler to apply confidence-based highlighting
        header_template: Optional Excel template to use
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True, parents=True)

    header_path = Path(header_template) if header_template else None

    digitization_files = discover_files(input_dir, "digitization.json")

    for digitization_file in tqdm(digitization_files, desc="Writing Excel files"):
        output_path = reroute_file(
            digitization_file, output_dir, input_dir
        ).with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        digitization_df = pl.read_json(digitization_file)
        wb = write_excel_workbook(digitization_df, aggregator, styler, header_path)

        wb.save(output_path)
        wb.close()


def main():
    parser = create_parser()
    args = parser.parse_args()
    input_dir = Path(args.input_dir)
    assert input_dir.exists(), f"Input directory {input_dir} does not exist"

    if args.output_dir is None:
        output_dir = input_dir
    else:
        output_dir = Path(args.output_dir)

    aggregator = AggregatorPipeline(
        [MostOccurringValues(at_least=2), HighestScoring()],
    )
    styler = NivoDefaultStyler(threshold=args.confidence_threshold or 0.9)

    write_excels_batch(
        input_dir=input_dir,
        output_dir=output_dir,
        aggregator=aggregator,
        styler=styler,
        header_template=args.header_template,
    )


if __name__ == "__main__":
    main()
