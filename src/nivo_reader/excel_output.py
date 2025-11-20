"""Excel output utilities for NIVO table data."""

from pathlib import Path
from typing import Any

import cv2
import openpyxl
import openpyxl.utils
from cv2.typing import MatLike
from openpyxl.styles import Alignment, Font, PatternFill


def write_tables_to_excel(
    cell_contents: list[Any],
    confidence_values: list[float | None],
    coordinates: list[tuple[int, int]],
    output_file: str,
    row_headers: list[str | None],
    row_headers_similarities: list[float | None],
    low_confidence_threshold: float = 0.7,
) -> None:
    """Write cell contents and confidence values to Excel file.

    Creates two sheets: "Cell Contents" and "Confidence Values".
    Marks low-confidence cells in red.

    Args:
        cell_contents: list of cell values
        confidence_values: list of confidence scores (0.0 to 1.0)
        coordinates: list of (row, col) tuples
        output_file: Output Excel file path
        row_headers: Station names for row headers
        row_headers_similarities: Similarity scores for station names
        low_confidence_threshold: Minimum confidence to display value

    Raises:
        ValueError: If input lists have different lengths or no data provided
    """
    # Validate input
    if not (len(cell_contents) == len(confidence_values) == len(coordinates)):
        raise ValueError(
            f"Input lists must have same length. Got: contents={len(cell_contents)}, "
            f"confidence={len(confidence_values)}, coordinates={len(coordinates)}"
        )

    if not coordinates:
        raise ValueError("No data provided")

    # Create workbook with two sheets
    wb = openpyxl.Workbook()

    # Remove default sheet and create named sheets
    if wb.active:
        wb.remove(wb.active)
    sheet_contents = wb.create_sheet("Cell Contents")
    sheet_confidence = wb.create_sheet("Confidence Values")

    # Determine table dimensions
    max_row = max(coord[0] for coord in coordinates) + 1
    max_col = max(coord[1] for coord in coordinates) + 1

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Write row headers if provided
    if row_headers:
        for row_idx, (header_text, similarity) in enumerate(
            zip(row_headers, row_headers_similarities)
        ):
            excel_row = row_idx + 1

            is_low_similarity = similarity < -2 if similarity else True

            # Write to first column
            header_cell = sheet_contents.cell(row=excel_row, column=1)
            header_cell.value = header_text
            header_cell.alignment = Alignment(horizontal="left", vertical="center")
            header_cell.font = Font(bold=True)
            if is_low_similarity:
                header_cell.fill = red_fill

            # Also write to confidence sheet
            conf_header_cell = sheet_confidence.cell(row=excel_row, column=1)
            conf_header_cell.value = similarity
            conf_header_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # Write data to both sheets
    for content, confidence, (row, col) in zip(
        cell_contents, confidence_values, coordinates
    ):
        # Excel uses 1-indexed cells
        excel_row = row + 1
        excel_col = col + 2

        # Check if confidence is below threshold
        is_low_confidence = (
            confidence < low_confidence_threshold if confidence else True
        )

        # Write content
        cell = sheet_contents.cell(row=excel_row, column=excel_col)
        cell.value = content
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_low_confidence:
            cell.value = "-"
            cell.fill = red_fill

        # Write confidence
        conf_cell = sheet_confidence.cell(row=excel_row, column=excel_col)
        conf_cell.value = confidence
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_low_confidence:
            conf_cell.fill = red_fill

        # Format confidence as percentage with 2 decimal places
        conf_cell.number_format = "0.00%"

    # Auto-adjust column widths
    for sheet in [sheet_contents, sheet_confidence]:
        for col in range(1, max_col + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col)

            for row in range(1, max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    Path(output_file).parent.mkdir(exist_ok=True, parents=True)
    wb.save(output_file)


def save_artifacts(
    artifacts: dict[str, MatLike],
    output_dir: Path | str,
) -> None:
    """Save image artifacts to directory.

    Args:
        artifacts: Dictionary mapping names to images
        output_dir: Output directory path
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for name, image in artifacts.items():
        cv2.imwrite(str(output_dir / f"{name}.jpg"), image)


def draw_bounding_boxes(
    base_image: MatLike,
    contours: list,
    boxes: bool,
    color: tuple[int, int, int] = (0, 255, 0),
    thickness: int = 5,
    width_factor: float = 0.0,
    height_factor: float = 0.0,
) -> MatLike:
    """Draw bounding boxes on image.

    Args:
        base_image: Base image to draw on
        contours: list of contours or boxes
        boxes: True if contours are already boxes, False if contours
        color: RGB color tuple
        thickness: Line thickness
        width_factor: Horizontal adjustment factor
        height_factor: Vertical adjustment factor

    Returns:
        Image with drawn boxes
    """
    overlay = base_image.copy()
    if len(overlay.shape) == 2:  # If grayscale, convert to RGB
        overlay = cv2.cvtColor(overlay, cv2.COLOR_GRAY2RGB)

    for contour in contours:
        if contour is not None and len(contour) > 0:
            if boxes:
                x, y, w, h = contour
            else:
                x, y, w, h = cv2.boundingRect(contour)

            # Adjust bounding box dimensions
            x += int(w * width_factor)
            y -= int(h * height_factor)
            w -= int(w * width_factor)
            h += int(h * height_factor * 2)

            cv2.rectangle(overlay, (x, y), (x + w, y + h), color, thickness)

    return overlay


def draw_straight_lines(
    base_image: MatLike,
    line_offsets: list,
    orientation: str,
    color: tuple[int, int, int] = (255, 0, 0),
    thickness: int = 1,
    **kwargs,
) -> MatLike:
    """Draw lines on image.

    Args:
        base_image: Base image to draw on
        line_offsets: list of line positions
        orientation: "horizontal" or "vertical"
        color: RGB color tuple
        thickness: Line thickness

    Returns:
        Image with drawn lines
    """
    if base_image.ndim == 2:
        base_image = cv2.cvtColor(base_image, cv2.COLOR_GRAY2BGR)
    overlay = base_image.copy()
    for line_offset in line_offsets:
        if orientation == "horizontal":
            pt1 = (0, line_offset)
            pt2 = (overlay.shape[1], line_offset)
        elif orientation == "vertical":
            pt1 = (line_offset, 0)
            pt2 = (line_offset, overlay.shape[0])
        else:
            raise ValueError(f"Invalid orientation: {orientation}")
        cv2.line(overlay, pt1, pt2, color, thickness, **kwargs)
    return overlay
